import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os

def parse_csv(filepath):
    with open(filepath, 'r', encoding='utf-8-sig') as f:
        raw = f.read()
    data = {}
    lines = [l.strip() for l in raw.replace('\r\n', '\n').split('\n')]
    for line in lines:
        if line.startswith('Date,'):
            data['date'] = line.split(',', 1)[1].strip()
    for i, line in enumerate(lines):
        if line.startswith(',Retail Net OPS'):
            data['sw'] = lines[i+1].split(',')
            data['pw'] = lines[i+2].split(',')
            data['py'] = lines[i+3].split(',')
            break
    daily = {}
    for i, line in enumerate(lines):
        if line.startswith('Period,Merchant Type,Metric Type,'):
            for j in range(i+1, len(lines)):
                r = lines[j]
                if not r or r.startswith('Product Group') or r.startswith('Top ASINs'):
                    break
                parts = r.split(',')
                if len(parts) >= 10:
                    key = f"{parts[0]}|{parts[1]}|{parts[2]}"
                    daily[key] = parts[3:]
            break
    data['daily'] = daily
    pg_ops, pg_units = {}, {}
    metric = None
    for line in lines:
        if line.startswith('Product Group breakdown,NET_OPS'):
            metric = 'ops'
            continue
        elif line.startswith('Product Group breakdown,NET_UNITS'):
            metric = 'units'
            continue
        elif line.startswith('Top ASINs'):
            metric = None
            continue
        if metric and line:
            parts = line.split(',', 1)
            if len(parts) == 2 and parts[0]:
                try:
                    if metric == 'ops':
                        pg_ops[parts[0]] = float(parts[1])
                    else:
                        pg_units[parts[0]] = int(float(parts[1]))
                except ValueError:
                    pass
    data['pg_ops'] = pg_ops
    data['pg_units'] = pg_units
    top_asins = []
    in_asins = False
    for line in lines:
        if line.startswith('Top ASINs,NET_OPS'):
            in_asins = True
            continue
        elif line.startswith('Top ASINs,NET_UNITS'):
            in_asins = False
            continue
        if in_asins and line and line[0] == 'B':
            parts = line.split(',', 2)
            if len(parts) >= 2:
                try:
                    top_asins.append({'asin': parts[0], 'value': float(parts[1]),
                                      'title': parts[2] if len(parts) > 2 else ''})
                except ValueError:
                    pass
    data['top_asins'] = top_asins
    return data

def fnum(v):
    try:
        return float(v)
    except Exception:
        return 0

w13 = parse_csv('W13/2026_AU30_w13.csv')
w12 = parse_csv('W13/2026_AU30_w12.csv')
ly  = parse_csv('W13/2025_AU30_w13.csv')

hdr_font = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
hdr_fill = PatternFill('solid', fgColor='4472C4')
title_font = Font(name='Calibri', bold=True, size=14)
sub_font = Font(name='Calibri', bold=True, size=12)
bold10 = Font(name='Calibri', bold=True, size=10)
norm10 = Font(name='Calibri', size=10)
gf = Font(name='Calibri', size=10, color='006100')
rf = Font(name='Calibri', size=10, color='9C0006')
gfill = PatternFill('solid', fgColor='C6EFCE')
rfill = PatternFill('solid', fgColor='FFC7CE')
gray = PatternFill('solid', fgColor='F2F2F2')
bdr = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
ca = Alignment(horizontal='center', vertical='center', wrap_text=True)
ra = Alignment(horizontal='right', vertical='center')
la = Alignment(horizontal='left', vertical='center', wrap_text=True)

def pct_cell(ws, r, c, val):
    cell = ws.cell(row=r, column=c, value=val)
    cell.number_format = '0.0%'
    cell.alignment = ra
    cell.border = bdr
    cell.fill = gfill if val >= 0 else rfill
    cell.font = gf if val >= 0 else rf

def hdr_row(ws, r, headers):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = ca
        cell.border = bdr

wb = openpyxl.Workbook()

# ── Sheet 1: Weekly Summary ─────────────────────────────────────
ws = wb.active
ws.title = 'Weekly Summary'
ws.merge_cells('A1:H1')
ws['A1'] = 'WBR - AU30 Pipeline Weekly Business Review'
ws['A1'].font = title_font
ws['A1'].alignment = ca
ws.merge_cells('A2:H2')
ws['A2'] = 'Week 13: 2026-03-22 ~ 2026-03-28 | amazon.com.au | USD'
ws['A2'].font = Font(name='Calibri', size=10, italic=True)
ws['A2'].alignment = ca

hdr_row(ws, 4, ['Metric', 'W13 2026', 'W12 2026', 'WoW Delta', 'WoW %', 'W13 2025', 'YoY Delta', 'YoY %'])

sw, pw, py_ = w13['sw'], w13['pw'], w13['py']
metrics = [
    ('3P Net OPS (USD)', 3, '#,##0.00'),
    ('3P Net Units', 4, '#,##0'),
    ('All Net OPS (USD)', 5, '#,##0.00'),
    ('All Net Units', 6, '#,##0'),
]
for i, (name, idx, nf) in enumerate(metrics):
    r = 5 + i
    s, p, y = fnum(sw[idx]), fnum(pw[idx]), fnum(py_[idx])
    ws.cell(row=r, column=1, value=name).font = bold10
    ws.cell(row=r, column=1).border = bdr
    for c, v in [(2, s), (3, p), (6, y)]:
        cell = ws.cell(row=r, column=c, value=v)
        cell.number_format = nf
        cell.alignment = ra
        cell.border = bdr
    wd = s - p
    ws.cell(row=r, column=4, value=wd).number_format = nf
    ws.cell(row=r, column=4).alignment = ra
    ws.cell(row=r, column=4).border = bdr
    wp = wd / p if p else 0
    pct_cell(ws, r, 5, wp)
    yd = s - y
    ws.cell(row=r, column=7, value=yd).number_format = nf
    ws.cell(row=r, column=7).alignment = ra
    ws.cell(row=r, column=7).border = bdr
    yp = yd / y if y else 0
    pct_cell(ws, r, 8, yp)

ws.column_dimensions['A'].width = 22
for c in range(2, 9):
    ws.column_dimensions[get_column_letter(c)].width = 18

# ── Sheet 2: Daily Trend ────────────────────────────────────────
ws2 = wb.create_sheet('Daily Trend')
ws2.merge_cells('A1:I1')
ws2['A1'] = 'Daily Net OPS & Units Trend - W13 2026'
ws2['A1'].font = title_font
ws2['A1'].alignment = ca

days_lbl = ['Sun 3/22', 'Mon 3/23', 'Tue 3/24', 'Wed 3/25', 'Thu 3/26', 'Fri 3/27', 'Sat 3/28']
d = w13['daily']

def daily_table(ws, start_row, title, key_suffix, nf, is_int=False):
    ws.cell(row=start_row, column=1, value=title).font = sub_font
    hr = start_row + 1
    hdr_row(ws, hr, ['Period'] + days_lbl + ['Total'])
    rows_data = [
        ('W13 2026', 'Selected week|All|' + key_suffix),
        ('W12 2026', 'Previous week|All|' + key_suffix),
        ('W13 2025', 'Previous year|All|' + key_suffix),
    ]
    for i, (label, key) in enumerate(rows_data):
        r = hr + 1 + i
        ws.cell(row=r, column=1, value=label).font = bold10
        ws.cell(row=r, column=1).border = bdr
        vals = d.get(key, ['0'] * 7)
        total = 0
        for j in range(7):
            v = int(fnum(vals[j])) if is_int else fnum(vals[j])
            total += v
            ws.cell(row=r, column=j + 2, value=v).number_format = nf
            ws.cell(row=r, column=j + 2).alignment = ra
            ws.cell(row=r, column=j + 2).border = bdr
        ws.cell(row=r, column=9, value=total).number_format = nf
        ws.cell(row=r, column=9).alignment = ra
        ws.cell(row=r, column=9).border = bdr
        ws.cell(row=r, column=9).font = bold10
    rw = hr + 4
    ws.cell(row=rw, column=1, value='WoW %').font = Font(name='Calibri', bold=True, size=10, italic=True)
    ws.cell(row=rw, column=1).border = bdr
    sv = d.get('Selected week|All|' + key_suffix, ['0'] * 7)
    pv = d.get('Previous week|All|' + key_suffix, ['0'] * 7)
    for j in range(7):
        s_val, p_val = fnum(sv[j]), fnum(pv[j])
        pct = (s_val - p_val) / p_val if p_val else 0
        pct_cell(ws, rw, j + 2, pct)
    ry = hr + 5
    ws.cell(row=ry, column=1, value='YoY %').font = Font(name='Calibri', bold=True, size=10, italic=True)
    ws.cell(row=ry, column=1).border = bdr
    yv = d.get('Previous year|All|' + key_suffix, ['0'] * 7)
    for j in range(7):
        s_val, y_val = fnum(sv[j]), fnum(yv[j])
        pct = (s_val - y_val) / y_val if y_val else 0
        pct_cell(ws, ry, j + 2, pct)
    return ry + 2

next_row = daily_table(ws2, 3, 'NET OPS (USD)', 'NET_OPS', '#,##0.00')
daily_table(ws2, next_row, 'NET UNITS', 'NET_UNITS', '#,##0', is_int=True)

ws2.column_dimensions['A'].width = 14
for c in range(2, 10):
    ws2.column_dimensions[get_column_letter(c)].width = 14

# ── Sheet 3: Product Group ──────────────────────────────────────
ws3 = wb.create_sheet('Product Group')
ws3.merge_cells('A1:F1')
ws3['A1'] = 'Product Group Breakdown - W13 2026'
ws3['A1'].font = title_font
ws3['A1'].alignment = ca

def pg_table(ws, start_row, title, d13, d12, dly, nf):
    ws.cell(row=start_row, column=1, value=title).font = sub_font
    hr = start_row + 1
    hdr_row(ws, hr, ['Product Group', 'W13 2026', 'W12 2026', 'W13 2025', 'WoW %', 'YoY %'])
    for i, (pg, val) in enumerate(d13.items()):
        r = hr + 1 + i
        ws.cell(row=r, column=1, value=pg).font = norm10
        ws.cell(row=r, column=1).border = bdr
        ws.cell(row=r, column=2, value=val).number_format = nf
        ws.cell(row=r, column=2).border = bdr
        v12 = d12.get(pg, 0)
        ws.cell(row=r, column=3, value=v12).number_format = nf
        ws.cell(row=r, column=3).border = bdr
        vly = dly.get(pg, 0)
        ws.cell(row=r, column=4, value=vly).number_format = nf
        ws.cell(row=r, column=4).border = bdr
        wow = (val - v12) / v12 if v12 else 0
        pct_cell(ws, r, 5, wow)
        yoy = (val - vly) / vly if vly else 0
        pct_cell(ws, r, 6, yoy)
    return hr + 1 + len(d13) + 1

nr = pg_table(ws3, 3, 'NET OPS by Product Group (USD)', w13['pg_ops'], w12['pg_ops'], ly['pg_ops'], '#,##0.00')
pg_table(ws3, nr, 'NET UNITS by Product Group', w13['pg_units'], w12['pg_units'], ly['pg_units'], '#,##0')

ws3.column_dimensions['A'].width = 22
for c in range(2, 7):
    ws3.column_dimensions[get_column_letter(c)].width = 16

# ── Sheet 4: Top ASINs ─────────────────────────────────────────
ws4 = wb.create_sheet('Top ASINs')
ws4.merge_cells('A1:E1')
ws4['A1'] = 'Top ASINs by Net OPS - W13 2026'
ws4['A1'].font = title_font
ws4['A1'].alignment = ca
hdr_row(ws4, 3, ['Rank', 'ASIN', 'Title', 'Net OPS (USD)', 'Share'])

asins = sorted(w13.get('top_asins', []), key=lambda x: x['value'], reverse=True)
total_ops = fnum(sw[5])
for i, a in enumerate(asins[:30]):
    r = 4 + i
    ws4.cell(row=r, column=1, value=i + 1).alignment = ca
    ws4.cell(row=r, column=1).border = bdr
    ws4.cell(row=r, column=2, value=a['asin']).border = bdr
    ws4.cell(row=r, column=3, value=a['title'][:80]).border = bdr
    ws4.cell(row=r, column=3).alignment = la
    ws4.cell(row=r, column=4, value=a['value']).number_format = '#,##0.00'
    ws4.cell(row=r, column=4).alignment = ra
    ws4.cell(row=r, column=4).border = bdr
    share = a['value'] / total_ops if total_ops else 0
    ws4.cell(row=r, column=5, value=share).number_format = '0.00%'
    ws4.cell(row=r, column=5).alignment = ra
    ws4.cell(row=r, column=5).border = bdr
    if i % 2 == 0:
        for cc in range(1, 6):
            ws4.cell(row=r, column=cc).fill = gray

ws4.column_dimensions['A'].width = 6
ws4.column_dimensions['B'].width = 14
ws4.column_dimensions['C'].width = 60
ws4.column_dimensions['D'].width = 18
ws4.column_dimensions['E'].width = 12

# ── Sheet 5: Key Highlights ────────────────────────────────────
ws5 = wb.create_sheet('Key Highlights')
ws5.merge_cells('A1:D1')
ws5['A1'] = 'W13 2026 Key Highlights'
ws5['A1'].font = title_font
ws5['A1'].alignment = ca

s_ops = fnum(sw[5])
p_ops = fnum(pw[5])
y_ops = fnum(py_[5])
s_units = fnum(sw[6])
p_units = fnum(pw[6])
y_units = fnum(py_[6])
wow_ops = (s_ops - p_ops) / p_ops if p_ops else 0
wow_units = (s_units - p_units) / p_units if p_units else 0
yoy_ops = (s_ops - y_ops) / y_ops if y_ops else 0
yoy_units = (s_units - y_units) / y_units if y_units else 0

def arrow(v):
    return '\u25b2' if v >= 0 else '\u25bc'

highlights = [
    'Net OPS WoW: {} {:.1f}% (${:,.0f} vs ${:,.0f})'.format(arrow(wow_ops), abs(wow_ops)*100, s_ops, p_ops),
    'Net Units WoW: {} {:.1f}% ({:,.0f} vs {:,.0f})'.format(arrow(wow_units), abs(wow_units)*100, s_units, p_units),
    'Net OPS YoY: {} {:.1f}% (${:,.0f} vs ${:,.0f})'.format(arrow(yoy_ops), abs(yoy_ops)*100, s_ops, y_ops),
    'Net Units YoY: {} {:.1f}% ({:,.0f} vs {:,.0f})'.format(arrow(yoy_units), abs(yoy_units)*100, s_units, y_units),
    '',
    'Top PG by OPS: ' + (list(w13['pg_ops'].keys())[0] if w13['pg_ops'] else 'N/A'),
    'Top PG by Units: ' + (list(w13['pg_units'].keys())[0] if w13['pg_units'] else 'N/A'),
    '',
    'Top ASIN: ' + (asins[0]['asin'] + ' - ' + asins[0]['title'][:60] if asins else 'N/A'),
]
for i, h in enumerate(highlights):
    ws5.cell(row=3 + i, column=1, value=h).font = Font(name='Calibri', size=11)
ws5.column_dimensions['A'].width = 90

# ── Save ────────────────────────────────────────────────────────
out = 'W13_WBR_AU30_Pipeline.xlsx'
wb.save(out)
print('WBR saved: ' + out + ' (' + str(os.path.getsize(out)) + ' bytes)')
